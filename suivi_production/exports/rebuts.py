from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.http import HttpResponse
from ..services.reporting import queryset_rebuts_par_of


def export_rebuts_pdf(numero: str = "", date_str: str = "") -> HttpResponse:
    ofs = list(queryset_rebuts_par_of(numero=numero, date_str=date_str))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rebuts_par_of.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("<b>Rebuts par Ordre de Fabrication</b>", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 0.2*cm))
    if numero or date_str:
        parts = []
        if numero:
            parts.append(f"Numéro contient '<i>{numero}</i>'")
        if date_str:
            parts.append(f"Date = <i>{date_str}</i>")
        story.append(Paragraph("Filtre: " + ", ".join(parts), styles['Normal']))
        story.append(Spacer(1, 0.3*cm))

    styleH = ParagraphStyle('HeaderCell', parent=styles['Normal'], fontSize=9, leading=11)
    styleCell = ParagraphStyle('BodyCell', parent=styles['Normal'], fontSize=9, leading=11)

    data = [[
        Paragraph('<b>Date</b>', styleH),
        Paragraph('<b>Numéro OF</b>', styleH),
        Paragraph('<b>Titre</b>', styleH),
        Paragraph('<b>Pièces finales</b>', styleH),
        Paragraph('<b>Qté à produire</b>', styleH),
        Paragraph('<b>Total rebuts</b>', styleH),
        Paragraph('<b>Taux (%)</b>', styleH),
    ]]
    total_prod = 0
    total_reb = 0
    for of in ofs:
        prod = int(of.quantite_produite_actuelle or 0)
        reb = int(of.total_rebut or 0)
        total = prod + reb
        taux = (reb / total * 100) if total > 0 else 0
        total_prod += prod
        total_reb += reb
        data.append([
            Paragraph(of.date_premiere_finalisation.strftime('%Y-%m-%d') if of.date_premiere_finalisation else '', styleCell),
            Paragraph(str(of.numero_of), styleCell),
            Paragraph(of.titre or '', styleCell),
            Paragraph(str(prod), styleCell),
            Paragraph(str(of.quantite_a_produire or 0), styleCell),
            Paragraph(str(reb), styleCell),
            Paragraph(f"{taux:.2f}", styleCell),
        ])

    total_decl = total_prod + total_reb
    taux_global = (total_reb / total_decl * 100) if total_decl > 0 else 0
    data.append([
        Paragraph('<b>TOTAL</b>', styleH),
        Paragraph('', styleH),
        Paragraph('', styleH),
        Paragraph(f'<b>{total_prod}</b>', styleH),
        Paragraph('', styleH),
        Paragraph(f'<b>{total_reb}</b>', styleH),
        Paragraph(f'<b>{taux_global:.2f}</b>', styleH),
    ])

    col_widths = [2.2*cm, 2.3*cm, 5.0*cm, 2.6*cm, 2.6*cm, 2.1*cm, 1.9*cm]
    tbl = Table(data, colWidths=col_widths, hAlign='LEFT')
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('ALIGN', (2,0), (-1,0), 'CENTER'),
        ('ALIGN', (2,1), (-1,-2), 'RIGHT'),
        ('ALIGN', (0,1), (1,-2), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.whitesmoke, colors.aliceblue]),
        ('LINEABOVE', (0,0), (-1,0), 1, colors.black),
        ('LINEBELOW', (0,0), (-1,0), 1, colors.black),
        ('LINEBELOW', (0,-1), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]))

    story.append(tbl)
    doc.build(story)
    return response


def export_rebuts_xlsx(numero: str = "", date_str: str = "") -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    ofs = list(queryset_rebuts_par_of(numero=numero, date_str=date_str))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rebuts par OF'

    headers = ['Numéro OF', 'Titre', 'Pièces finales', 'Quantité à produire', 'Total rebuts', 'Taux de rebut (%)']
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers)+1):
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')

    total_prod = 0
    total_reb = 0
    for of in ofs:
        prod = int(of.quantite_produite_actuelle or 0)
        reb = int(of.total_rebut or 0)
        total = prod + reb
        taux = (reb / total * 100) if total > 0 else 0
        total_prod += prod
        total_reb += reb
        ws.append([of.numero_of, of.titre or '', prod, of.quantite_a_produire or 0, reb, round(taux, 2)])

    total_decl = total_prod + total_reb
    taux_global = (total_reb / total_decl * 100) if total_decl > 0 else 0
    ws.append([])
    ws.append(['TOTAL', '', total_prod, '', total_reb, round(taux_global, 2)])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.cell(row=ws.max_row, column=3).font = bold
    ws.cell(row=ws.max_row, column=5).font = bold
    ws.cell(row=ws.max_row, column=6).font = bold

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_length + 2), 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rebuts_par_of.xlsx"'
    wb.save(response)
    return response
